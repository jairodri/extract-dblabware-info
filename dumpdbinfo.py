import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, colors
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime, date


def dump_dbinfo_to_csv(service_name:str, table_dataframes: dict, output_dir: str, sep: str=','):
    """
    Saves each DataFrame in the provided dictionary to a CSV file, organizing the files within a directory 
    named after the service.

    The CSV files are named after the corresponding table names. Each DataFrame is saved to a separate CSV file 
    in the specified output directory. If the directory does not include a subdirectory named after the service, 
    one will be created.

    Parameters:
    -----------
    service_name : str
        The name of the service (typically the database service) used to name the output subdirectory.
    
    table_dataframes : dict
        A dictionary where each key is a table name and each value is a DataFrame containing the table's data.
    
    output_dir : str
        The directory where the CSV files will be saved. If this directory does not contain a subdirectory named 
        after the service, one will be created.

    sep : str, optional
        Field delimiter for the output CSV files. The default is a comma.

    Returns:
    --------
    None
    """

    # Ensure the output directory includes a subdirectory named after the service
    if not output_dir.endswith(service_name):
        output_dir = os.path.join(output_dir, service_name)

    # Create the output directory if it does not exist
    os.makedirs(output_dir, exist_ok=True)

    # Iterate over each table name and its corresponding DataFrame in the dictionary
    for item, item_data in table_dataframes.items():
        table_name = item_data['name']
        dataframe = item_data['data']
        # Replace newline characters with spaces in all text columns
        dataframe = dataframe.applymap(lambda x: x.replace('\n', ' ') if isinstance(x, str) else x)

        # Create the CSV file path using the table name
        file_path = os.path.join(output_dir, f"{table_name}.csv")
        
        # Save the DataFrame to a CSV file with the specified delimiter and without the index
        dataframe.to_csv(file_path, sep=sep, index=False)


def create_hyperlink(ws, at_cell, sheet_name, cell_ref='A1', display_name=None, font_size=11):
    """
    Creates a hyperlink in a specified cell that links to another cell within the same workbook.

    This function adds a hyperlink to a specified cell in an Excel worksheet (`ws`). The hyperlink points 
    to a cell within another sheet (or the same sheet) within the same workbook. The cell containing the 
    hyperlink is formatted with a blue, underlined font to resemble a standard hyperlink.

    Parameters:
    -----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where the hyperlink will be created.
    
    at_cell : str
        The cell reference (e.g., 'B2') where the hyperlink will be placed in the `ws` worksheet.
    
    sheet_name : str
        The name of the sheet to which the hyperlink will point.
    
    cell_ref : str, optional
        The cell reference within the `sheet_name` sheet that the hyperlink will point to. Default is 'A1'.
    
    display_name : str, optional
        The text to be displayed in the cell containing the hyperlink. If not provided, defaults to the `sheet_name`.

    font_size : int, optional
        The font size to be applied to the cell containing the hyperlink. Default is 11.

    Returns:
    --------
    None
    """
    if display_name is None:
        display_name = sheet_name
    to_location = "'{0}'!{1}".format(sheet_name, cell_ref)
    ws[at_cell].hyperlink = Hyperlink(display=display_name, ref=at_cell, location=to_location)
    ws[at_cell].value = display_name
    ws[at_cell].font = Font(u='single', color=colors.BLUE, size=font_size)


def adjust_column_widths(sheet,  max_width=80):
    """
    Adjusts the width of each column in the Excel sheet based on the maximum width of the data and header values.

    Parameters:
    -----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where column widths need to be adjusted.

    max_width : int, optional (default=80)
        The maximum allowed width for any column. If the calculated width exceeds this value,
        the column width will be set to this maximum value.
    
    Returns:
    --------
    None
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Calculate the width required by the header (considering formatting)
        header_length = len(str(col[0].value))
        adjusted_header_length = header_length * 1.5  # Factor to account for bold and larger font size

        # Compare the header length with the lengths of the data values
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Use the greater of the header length or data length for column width
        max_length = max(max_length, adjusted_header_length)

        # Adjust the column width and apply the max_width limit
        adjusted_width = min(max_length + 2, max_width)
        sheet.column_dimensions[column].width = adjusted_width


def format_header_cell(cell, font_size=11):
    """
    Formats a header cell with the default styling: white bold text and green background.

    Parameters:
    -----------
    cell : openpyxl.cell.cell.Cell
        The cell to format.
    
    font_size : int, optional
        The font size to be applied to the header cell. Default is 11.

    Returns:
    --------
    None
    """
    cell.font = Font(color="FFFFFF", bold=True, size=font_size + 1)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")


def dump_dbinfo_to_excel(service_name:str, table_dataframes: dict, output_dir: str, include_record_count: bool = False, max_records_per_table: int = 50000):
    """
    Exports data from the provided dictionary to an Excel workbook, with each table's data in a separate sheet.

    This function creates an Excel workbook where each table's data is stored in its own sheet. The first sheet, titled 
    "Tables," serves as an index listing all table names, with hyperlinks to their respective sheets for easy navigation.

    Optionally, a column can be added to the index sheet that shows the number of records in each table.

    Each table's sheet will include a hyperlink in the header row to return to the index sheet. Columns of type CLOB are
    handled specially by saving their content to individual text files in a subdirectory and creating a hyperlink in the
    corresponding cell to access the text file directly.

    Parameters:
    -----------
    service_name : str
        The name of the database service, used to name the output Excel file and the directory where it will be saved.

    table_dataframes : dict
        A dictionary where each key is a table identifier, and each value is a dictionary containing:
            - 'name': The name of the table.
            - 'data': A pandas DataFrame containing the table's data.
            - 'fields': A dictionary describing each column's metadata including data type.

    output_dir : str
        The directory where the Excel file will be saved. The function ensures the directory structure includes a 
        folder named after the database service if it does not already exist.

    include_record_count : bool, optional
        If True, adds a column to the index sheet that shows the number of records in each table. Default is False.

    max_records_per_table : int, optional
        The maximum number of records to include per table in the Excel sheet. Default is 50,000.

    Returns:
    --------
    None
        The function saves the Excel workbook to the specified output directory and does not return any value.

    Notes:
    ------
    - The function ensures that the output directory exists, creating it if necessary.
    - Hyperlinks are created on the index sheet for easy navigation to each table's sheet, and each table's sheet 
      includes a hyperlink to return to the index sheet.
    - The number of records per table in the Excel file is limited by `max_records_per_table` to prevent excessive file size.
    - If `include_record_count` is set to True, the index sheet will include an additional column showing the number of records in each table.
    - For columns with data types 'CLOB' or 'LONG', the content is saved in separate text files in a 'CLOB' subdirectory, 
      and the cell in Excel contains a hyperlink to these files, allowing easy access to large text data.
    """

    # Ensure the output directory includes the database name
    if not output_dir.endswith(service_name):
        output_dir = os.path.join(output_dir, service_name)
    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the path for the CLOB subdirectory
    clob_subdir = os.path.join(output_dir, 'CLOB')
    # Ensure the CLOB subdirectory exists
    os.makedirs(clob_subdir, exist_ok=True)

    # Create the Excel workbook
    workbook = Workbook()

    # Default Excel font size if not specified
    standard_font_size = 11  

    # Use the default sheet as the index sheet
    index_sheet = workbook.active
    index_sheet.title = "Tables"
    
    # Add headers to the index sheet
    index_sheet.cell(row=1, column=1, value="Table")
    format_header_cell(index_sheet.cell(row=1, column=1), font_size=standard_font_size)
    if include_record_count:
        index_sheet.cell(row=1, column=2, value="Record Count")
        format_header_cell(index_sheet.cell(row=1, column=2), font_size=standard_font_size)

    # Populate the index sheet with table names (and record counts if enabled)
    for i, (item, item_data) in enumerate(table_dataframes.items(), start=2):
        table_name = item_data['name']
        dataframe = item_data['data']
        index_sheet.cell(row=i, column=1, value=table_name)
        if include_record_count:
            index_sheet.cell(row=i, column=2, value=len(dataframe))
    
    # Adjust column width for the index sheet
    adjust_column_widths(index_sheet)

    # Apply a filter to all columns
    index_sheet.auto_filter.ref = index_sheet.dimensions

    for item, item_data in table_dataframes.items():
        table_name = item_data['name']
        dataframe = item_data['data']
        fields = item_data['fields']
        column_names = list(fields.keys())
        data_types = [value['data_type'] for value in fields.values()]
        # Limit the number of records to max_records_per_table
        limited_dataframe = dataframe.head(max_records_per_table)

        # Create a new sheet with the table name
        sheet = workbook.create_sheet(title=table_name)
        
        # Freeze first row (header)
        sheet.freeze_panes = 'A2'  

        # Add the DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(limited_dataframe, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Apply formatting to header row
                if r_idx == 1:  
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    format_header_cell(cell, font_size=standard_font_size)
                else:
                    if data_types[c_idx-1] == 'LONG' or data_types[c_idx-1] == 'CLOB':
                        # Handle CLOB data by writing it to a text file
                        if pd.notna(value):
                            # Define the filename for the CLOB content
                            clob_filename = f"{table_name}_{column_names[c_idx-1]}_{r_idx:06}.txt"
                            clob_filepath = os.path.join(clob_subdir, clob_filename)
                            
                            # Write CLOB content to a text file
                            with open(clob_filepath, 'w', encoding='utf-8') as file:
                                file.write(str(value))
                            
                            # Set the cell value to the filename
                            cell_value = clob_filename

                            # Create a hyperlink in the Excel cell to the CLOB text file
                            cell = sheet.cell(row=r_idx, column=c_idx, value=clob_filename)
                            cell.hyperlink = clob_filepath
                            cell.style = "Hyperlink"

                        else:
                            # In case of a NaN CLOB value
                            cell_value = ''
                    # Convert datetime64 and date values to Excel date format
                    elif isinstance(value, pd.Timestamp):
                        cell_value = value.to_pydatetime()
                    elif isinstance(value, date):
                        cell_value = value.date()  
                    else:
                        # Ensure the value is converted to a string if it's not a basic data type
                        cell_value = str(value) if not isinstance(value, (int, float, type(None))) else value
                    cell = sheet.cell(row=r_idx, column=c_idx, value=cell_value)

        
        # Auto-size columns 
        adjust_column_widths(sheet)

        # Apply a filter to all columns
        sheet.auto_filter.ref = sheet.dimensions

        # Add a hyperlink to return to the "Tables" sheet in the last cell of the header row
        last_col_idx = len(dataframe.columns) + 1
        return_cell = sheet.cell(row=1, column=last_col_idx)
        create_hyperlink(sheet, at_cell=return_cell.coordinate, sheet_name="Tables", cell_ref='A1', display_name="Return to Tables", font_size=standard_font_size+1)

        # Adjust the column width to fit the "Return to Tables" message.
        sheet.column_dimensions[return_cell.column_letter].width = len("Return to Tables") + 2


    # Links are created for each table in the list for easy access to its sheet.
    for i in range(2, index_sheet.max_row + 1):
        create_hyperlink(index_sheet, 'A' + str(i), index_sheet['A' + str(i)].value, cell_ref='A1')

    # Save the workbook to the output directory with the name of the database
    excel_file_path = os.path.join(output_dir, f"{service_name}.xlsx")
    workbook.save(excel_file_path)
