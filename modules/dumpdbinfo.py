"""
Database Information Export Module

This module provides functionality to export Oracle database information
to CSV and Excel formats with comprehensive formatting and navigation features.

Author: Database Analysis Team
Version: 2.0
"""

import os
import re
import shutil
from datetime import datetime, date
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils.exceptions import IllegalCharacterError

from modules.utils import adjust_column_widths, format_header_cell


# ==============================================================================
# UTILITY FUNCTIONS
# ==============================================================================

def clean_value(value: Any) -> Any:
    """
    Clean values by removing illegal characters that cannot be written to Excel.
    
    Removes ASCII control characters (0-31) except newline and tab while preserving
    other printable characters. Handles None values by converting to empty string.

    Args:
        value (Any): Value to be cleaned

    Returns:
        Any: Cleaned value safe for Excel export
    """
    if value is None:
        return ""

    if isinstance(value, str):
        # Remove invalid control characters (ASCII < 32) except '\n' and '\t'
        value = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', value)
    
    return value


def create_hyperlink(worksheet, at_cell: str, sheet_name: str, cell_ref: str = 'A1', 
                    display_name: Optional[str] = None, font_size: int = 11) -> None:
    """
    Create a hyperlink in a specified cell that links to another sheet within the same workbook.

    Args:
        worksheet: The openpyxl worksheet where the hyperlink will be created
        at_cell (str): Cell reference (e.g., 'B2') where the hyperlink will be placed
        sheet_name (str): Name of the target sheet for the hyperlink
        cell_ref (str): Target cell reference within the sheet (default: 'A1')
        display_name (str, optional): Text to display in the cell (default: sheet_name)
        font_size (int): Font size for the hyperlink (default: 11)

    Returns:
        None
    """
    if display_name is None:
        display_name = sheet_name
        
    to_location = f"'{sheet_name}'!{cell_ref}"
    worksheet[at_cell].hyperlink = Hyperlink(display=display_name, ref=at_cell, location=to_location)
    worksheet[at_cell].value = display_name
    worksheet[at_cell].font = Font(underline='single', color='0000FF', size=font_size)


def create_output_directory(output_dir: str, folder_name: str) -> str:
    """
    Create output directory structure for data export.
    
    Args:
        output_dir (str): Base output directory path
        folder_name (str): Subfolder name to create
        
    Returns:
        str: Full path to the created directory
    """
    # Ensure the output directory includes the specified folder
    if not output_dir.endswith(folder_name):
        output_dir = os.path.join(output_dir, folder_name)

    # Delete and recreate the output directory for clean export
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)

    os.makedirs(output_dir, exist_ok=True)
    print(f"Created output directory: {output_dir}")
    
    return output_dir


# ==============================================================================
# CSV EXPORT FUNCTIONS
# ==============================================================================

def dump_dbinfo_to_csv(folder_name: str, table_dataframes: Dict[str, Any], 
                      output_dir: str, sep: str = ',', suffix: Optional[str] = None) -> None:
    """
    Export database table information to CSV files.

    Creates separate CSV files for each table in the provided dictionary,
    organizing them within a directory named after the database service.

    Args:
        folder_name (str): Name of the output subdirectory
        table_dataframes (Dict[str, Any]): Dictionary where keys are table identifiers
                                          and values contain table metadata and data
        output_dir (str): Base directory for CSV file output
        sep (str): Field delimiter for CSV files (default: ',')
        suffix (str, optional): Suffix to append to filenames

    Returns:
        None

    Notes:
        - Replaces newline characters with spaces in text columns for CSV compatibility
        - Creates one CSV file per table with clean, structured data
        - Automatically handles CLOB, LONG, VARCHAR, and VARCHAR2 data types
    """
    print('Exporting database information to CSV files...')
    
    # Create output directory structure
    output_dir = create_output_directory(output_dir, folder_name)

    # Process each table in the dictionary
    for item, item_data in table_dataframes.items():
        table_name = item_data['name']
        dataframe = item_data['data'].copy()  # Work with a copy to avoid modifying original
        fields = item_data['fields']

        # Clean text columns by replacing newline characters with spaces
        for column_name, column_info in fields.items():
            if column_info['data_type'] in ['CLOB', 'LONG', 'VARCHAR', 'VARCHAR2']:
                try:
                    # Column names in dataframe are lowercase, while in fields dictionary are uppercase
                    if column_name.lower() in dataframe.columns:
                        dataframe[column_name.lower()] = dataframe[column_name.lower()].astype(str).str.replace(r'\r?\n', ' ', regex=True)
                except KeyError:
                    print(f'Warning: Column {column_name} not found in table {table_name}')

        # Generate output filename
        output_table_name = f"{table_name}{suffix}" if suffix else table_name
        file_path = os.path.join(output_dir, f"{output_table_name}.csv")
        
        # Export DataFrame to CSV
        dataframe.to_csv(file_path, sep=sep, index=False, encoding='utf-8')
        print(f"Exported {len(dataframe)} records to {output_table_name}.csv")


# ==============================================================================
# EXCEL EXPORT FUNCTIONS
# ==============================================================================

def _process_clob_data(value: Any, table_name: str, column_name: str, index_values: List[str], 
                      clob_subdir: str) -> str:
    """
    Process CLOB data by saving it to a text file and returning the filename.
    
    Args:
        value: CLOB value to process
        table_name (str): Name of the table containing the CLOB
        column_name (str): Name of the CLOB column
        index_values (List[str]): Values to use for unique filename generation
        clob_subdir (str): Directory path for CLOB text files
        
    Returns:
        str: Filename of the created text file
    """
    if pd.notna(value):
        # Build unique filename using table, column, and index values
        index_part = "_".join(str(val) for val in index_values)
        clob_filename = f"{table_name}__{column_name}_{index_part}.txt"
        
        # Replace invalid characters with underscores
        clob_filename = re.sub(r'[^\w_. -]', '_', clob_filename)
        clob_filepath = os.path.join(clob_subdir, clob_filename)
        
        # Write CLOB content to text file
        try:
            with open(clob_filepath, 'w', encoding='utf-8') as file:
                file.write(str(value))
            return clob_filename
        except Exception as e:
            print(f"Error writing CLOB file {clob_filename}: {e}")
            return "CLOB_ERROR"
    else:
        return ""


def dump_dbinfo_to_excel(folder_name: str, table_dataframes: Dict[str, Any], output_dir: str, 
                        include_record_count: bool = False, max_records_per_table: int = 50000, 
                        file_name: Optional[str] = None) -> None:
    """
    Export database table information to Excel workbook with enhanced formatting and navigation.

    Creates an Excel workbook with an index sheet and separate sheets for each table.
    Includes hyperlinks for navigation, special handling for CLOB data, and professional formatting.

    Args:
        folder_name (str): Used for naming the output Excel file and directory
        table_dataframes (Dict[str, Any]): Dictionary containing table metadata and data
        output_dir (str): Directory where the Excel file will be saved
        include_record_count (bool): Whether to include record counts in index sheet
        max_records_per_table (int): Maximum records to include per table sheet
        file_name (str, optional): Custom name for the Excel file

    Returns:
        None

    Features:
        - Creates navigable index sheet with hyperlinks to all table sheets
        - Handles CLOB data by saving to external text files with hyperlinks
        - Applies professional formatting with frozen headers and auto-filters
        - Automatically adjusts column widths for optimal readability
        - Limits records per sheet to prevent excessive file sizes
    """
    print('Exporting database information to Excel workbook...')
    
    # Create output directory structure
    output_dir = create_output_directory(output_dir, folder_name)

    # Create CLOB subdirectory for large text data
    clob_subdir = os.path.join(output_dir, 'CLOB')
    os.makedirs(clob_subdir, exist_ok=True)

    # Initialize Excel workbook
    workbook = Workbook()
    standard_font_size = 11

    # Configure index sheet
    index_sheet = workbook.active
    index_sheet.title = "Tables"
    
    # Set up index sheet headers
    headers = ["Table"]
    if include_record_count:
        headers.append("Record Count")
        
    for col_idx, header in enumerate(headers, 1):
        cell = index_sheet.cell(row=1, column=col_idx, value=header)
        format_header_cell(cell, font_size=standard_font_size)

    # Populate index sheet with table information
    for row_idx, (item, item_data) in enumerate(table_dataframes.items(), start=2):
        table_name = item_data['name']
        dataframe = item_data['data']
        
        # Add table name
        index_sheet.cell(row=row_idx, column=1, value=table_name)
        
        # Add record count if requested
        if include_record_count:
            index_sheet.cell(row=row_idx, column=2, value=len(dataframe))

    # Apply formatting to index sheet
    adjust_column_widths(index_sheet)
    index_sheet.auto_filter.ref = index_sheet.dimensions

    # Create individual sheets for each table
    for item, item_data in table_dataframes.items():
        table_name = item_data['name']
        dataframe = item_data['data']
        fields = item_data['fields']
        index_list = item_data.get('index', [])

        # Limit records to prevent oversized files
        limited_dataframe = dataframe.head(max_records_per_table)

        # Create new sheet for the table
        sheet = workbook.create_sheet(title=table_name)
        sheet.freeze_panes = 'A2'  # Freeze header row

        # Get column information for data type processing
        column_names = list(fields.keys())
        data_types = [field_info['data_type'] for field_info in fields.values()]

        # Populate sheet with data
        for r_idx, row in enumerate(dataframe_to_rows(limited_dataframe, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if r_idx == 1:  # Header row
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    format_header_cell(cell, font_size=standard_font_size)
                else:  # Data rows
                    # Handle CLOB/LONG data types specially
                    if c_idx <= len(data_types) and data_types[c_idx-1] in ['LONG', 'CLOB']:
                        if index_list:
                            # Generate unique identifier for CLOB file
                            index_values = [
                                str(sheet.cell(row=r_idx, column=column_names.index(index_col)+1).value) 
                                for index_col in index_list if index_col in column_names
                            ]
                            
                            clob_filename = _process_clob_data(
                                value, table_name, column_names[c_idx-1], index_values, clob_subdir
                            )
                            
                            # Create hyperlink to CLOB file
                            if clob_filename:
                                cell = sheet.cell(row=r_idx, column=c_idx, value=clob_filename)
                                if clob_filename != "CLOB_ERROR":
                                    cell.hyperlink = os.path.join(clob_subdir, clob_filename)
                                    cell.style = "Hyperlink"
                        else:
                            cell = sheet.cell(row=r_idx, column=c_idx, value="CLOB_NO_INDEX")
                    
                    # Handle datetime objects
                    elif isinstance(value, pd.Timestamp):
                        cell_value = value.to_pydatetime()
                        cell = sheet.cell(row=r_idx, column=c_idx, value=cell_value)
                    elif isinstance(value, date):
                        cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Handle other data types
                    else:
                        cell_value = clean_value(value)
                        try:
                            cell = sheet.cell(row=r_idx, column=c_idx, value=cell_value)
                        except IllegalCharacterError:
                            print(f'Illegal character found in table {table_name}, replacing with cleaned value')
                            cell = sheet.cell(row=r_idx, column=c_idx, value=str(cell_value).encode('ascii', 'ignore').decode('ascii'))

        # Apply sheet formatting
        adjust_column_widths(sheet)
        sheet.auto_filter.ref = sheet.dimensions

        # Add return link to index
        last_col_idx = len(dataframe.columns) + 1
        return_cell = sheet.cell(row=1, column=last_col_idx)
        create_hyperlink(sheet, return_cell.coordinate, "Tables", 'A1', 
                        "Return to Tables", standard_font_size + 1)
        sheet.column_dimensions[return_cell.column_letter].width = len("Return to Tables") + 2

    # Create hyperlinks from index to individual sheets
    for i in range(2, index_sheet.max_row + 1):
        table_name = index_sheet[f'A{i}'].value
        if table_name:
            create_hyperlink(index_sheet, f'A{i}', table_name, 'A1')

    # Save the workbook
    output_filename = file_name if file_name else folder_name
    excel_file_path = os.path.join(output_dir, f"{output_filename}.xlsx")
    
    try:
        workbook.save(excel_file_path)
        print(f"Excel workbook saved successfully: {excel_file_path}")
        print(f"Created {len(table_dataframes)} table sheets with up to {max_records_per_table:,} records each")
    except Exception as e:
        print(f"Error saving Excel workbook: {e}")


# ==============================================================================
# SCHEMA COMPARISON REPORT FUNCTIONS
# ==============================================================================

def generate_unified_schema_report(schema_repository: Dict[str, pd.DataFrame], 
                                 diff_df: pd.DataFrame, output_filename: str) -> None:
    """
    Generate a unified Excel report containing schema information and comparison results.

    Creates a comprehensive Excel workbook with:
    1. Index sheet with schema summary and navigation links
    2. Individual sheets for each schema structure
    3. Differences analysis sheet with color-coded results

    Args:
        schema_repository (Dict[str, pd.DataFrame]): Dictionary of schema DataFrames
        diff_df (pd.DataFrame): DataFrame containing schema differences
        output_filename (str): Name of the output Excel file

    Returns:
        None

    Features:
        - Professional navigation with hyperlinks between sheets
        - Color-coded difference analysis for easy identification
        - Automatic column width adjustment and filtering
        - Comprehensive validation and error handling
    """
    print(f"Generating unified schema report: '{output_filename}'...")
    
    try:
        # Initial validation
        if not schema_repository:
            print("Error: Schema repository is empty")
            return
        
        # Validate and normalize schemas
        validated_schemas = {}
        for schema_name, df in schema_repository.items():
            if df is None or df.empty:
                print(f"Warning: Schema '{schema_name}' is empty, skipping")
                continue
            
            # Normalize column names to uppercase
            df_normalized = df.copy()
            df_normalized.columns = df_normalized.columns.str.upper()
            
            # Verify required columns exist
            required_columns = ['TABLE_NAME', 'COLUMN_NAME']
            missing_columns = [col for col in required_columns if col not in df_normalized.columns]
            
            if missing_columns:
                print(f"Error: Schema '{schema_name}' missing required columns: {missing_columns}")
                print(f"Available columns: {list(df_normalized.columns)}")
                continue
            
            validated_schemas[schema_name] = df_normalized
        
        if not validated_schemas:
            print("Error: No valid schemas to process")
            return
        
        # Initialize workbook
        workbook = Workbook()
        standard_font_size = 11
        
        # Create index sheet
        index_sheet = workbook.active
        index_sheet.title = "Schema_Index"
        
        # Configure index sheet headers
        headers = ["Schema", "Table Count", "Column Count"]
        for col_idx, header in enumerate(headers, 1):
            cell = index_sheet.cell(row=1, column=col_idx, value=header)
            format_header_cell(cell, font_size=standard_font_size)
        
        # Populate index sheet with schema information
        for row_idx, (schema_name, df) in enumerate(validated_schemas.items(), start=2):
            # Create hyperlink to schema sheet
            cell = index_sheet.cell(row=row_idx, column=1, value=schema_name)
            create_hyperlink(index_sheet, cell.coordinate, schema_name, 'A1', 
                           schema_name, standard_font_size)
            
            # Calculate and add table count
            try:
                num_tables = df['TABLE_NAME'].nunique() if 'TABLE_NAME' in df.columns else 0
                index_sheet.cell(row=row_idx, column=2, value=num_tables)
            except Exception as e:
                print(f"Error calculating table count for {schema_name}: {e}")
                index_sheet.cell(row=row_idx, column=2, value="Error")
            
            # Add total column count
            index_sheet.cell(row=row_idx, column=3, value=len(df))
        
        # Add differences section to index
        diff_row = len(validated_schemas) + 3
        index_sheet.cell(row=diff_row, column=1, value="--- DIFFERENCE ANALYSIS ---")
        format_header_cell(index_sheet.cell(row=diff_row, column=1), font_size=standard_font_size)
        
        diff_row += 1
        cell = index_sheet.cell(row=diff_row, column=1, value="Schema Differences")
        create_hyperlink(index_sheet, cell.coordinate, "Schema_Differences", 'A1', 
                        "View Differences", standard_font_size)
        
        # Add difference statistics
        if not diff_df.empty:
            index_sheet.cell(row=diff_row, column=2, value=len(diff_df))
            index_sheet.cell(row=diff_row, column=3, value="differences found")
        else:
            index_sheet.cell(row=diff_row, column=2, value="0")
            index_sheet.cell(row=diff_row, column=3, value="Schemas are identical!")
        
        # Format index sheet
        adjust_column_widths(index_sheet)
        index_sheet.auto_filter.ref = index_sheet.dimensions
        
        # Create individual schema sheets
        for schema_name, df in validated_schemas.items():
            try:
                sheet = workbook.create_sheet(title=schema_name)
                
                # Add data with header formatting
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:  # Header row
                            format_header_cell(cell, font_size=standard_font_size)
                
                # Format sheet
                adjust_column_widths(sheet)
                sheet.auto_filter.ref = sheet.dimensions
                sheet.freeze_panes = 'A2'
                
                # Add return link
                last_col_idx = len(df.columns) + 1
                return_cell = sheet.cell(row=1, column=last_col_idx)
                create_hyperlink(sheet, return_cell.coordinate, "Schema_Index", 'A1', 
                               "Back to Index", standard_font_size+1)
                sheet.column_dimensions[return_cell.column_letter].width = len("Back to Index") + 2
                
            except Exception as e:
                print(f"Error creating sheet for schema '{schema_name}': {e}")
                continue
        
        # Create differences sheet
        if not diff_df.empty:
            try:
                diff_sheet = workbook.create_sheet(title="Schema_Differences")
                
                # Add differences data
                for r_idx, row in enumerate(dataframe_to_rows(diff_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = diff_sheet.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:  # Header row
                            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF', size=standard_font_size+1)
                
                # Apply color coding for different types of differences
                difference_colors = {
                    'Table Missing': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),
                    'Column Missing': PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid'),
                }
                
                # Color-code difference rows
                for row_num in range(2, len(diff_df) + 2):
                    diff_type_cell = diff_sheet.cell(row=row_num, column=3)
                    diff_type = diff_type_cell.value
                    
                    # Determine appropriate color
                    if any(keyword in str(diff_type) for keyword in ['Different', 'different']):
                        fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                    else:
                        fill = difference_colors.get(diff_type, None)
                    
                    # Apply color to entire row
                    if fill:
                        for col_num in range(1, len(diff_df.columns) + 1):
                            diff_sheet.cell(row=row_num, column=col_num).fill = fill
                
                # Format differences sheet
                adjust_column_widths(diff_sheet)
                diff_sheet.auto_filter.ref = diff_sheet.dimensions
                diff_sheet.freeze_panes = 'D2'
                
                # Add return link
                last_col_idx = len(diff_df.columns) + 1
                return_cell = diff_sheet.cell(row=1, column=last_col_idx)
                create_hyperlink(diff_sheet, return_cell.coordinate, "Schema_Index", 'A1', 
                               "Back to Index", standard_font_size+1)
                diff_sheet.column_dimensions[return_cell.column_letter].width = len("Back to Index") + 2
                
            except Exception as e:
                print(f"Error creating differences sheet: {e}")
        
        # Save workbook
        workbook.save(output_filename)
        print(f"Unified schema report generated successfully: '{output_filename}'")
        
        # Display summary
        print(f"Report summary:")
        print(f"- {len(validated_schemas)} schemas analyzed")
        print(f"- {len(diff_df) if not diff_df.empty else 0} differences found")
        print(f"- Sheets created: Index + {len(validated_schemas)} schemas + {'1' if not diff_df.empty else '0'} differences")
        
    except Exception as e:
        print(f"Error generating unified schema report '{output_filename}': {e}")
        import traceback
        traceback.print_exc()