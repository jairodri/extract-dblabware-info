from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


def adjust_column_widths(sheet,  max_width=80):
    """
    Adjusts the width of each column in the Excel sheet based on the maximum width of the data and header values.

    Parameters:
    -----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where column widths need to be adjusted.

    max_width : int, optional (default=80)
        The maximum allowed width for any column. If the calculated width exceeds this value,
        the column width will be set to this maximum value.
    
    Returns:
    --------
    None
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Calculate the width required by the header (considering formatting)
        header_length = len(str(col[0].value))
        adjusted_header_length = header_length * 1.5  # Factor to account for bold and larger font size

        # Compare the header length with the lengths of the data values
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Use the greater of the header length or data length for column width
        max_length = max(max_length, adjusted_header_length)

        # Adjust the column width and apply the max_width limit
        adjusted_width = min(max_length + 2, max_width)
        sheet.column_dimensions[column].width = adjusted_width


def format_header_cell(cell, font_size=11):
    """
    Formats a header cell with the default styling: white bold text and green background.

    Parameters:
    -----------
    cell : openpyxl.cell.cell.Cell
        The cell to format.
    
    font_size : int, optional
        The font size to be applied to the header cell. Default is 11.

    Returns:
    --------
    None
    """
    cell.font = Font(color="FFFFFF", bold=True, size=font_size + 1)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")