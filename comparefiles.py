import os
import difflib
import pandas as pd
from openpyxl import Workbook
import re
from dumpdbinfo import format_header_cell, adjust_column_widths
from datetime import datetime


def compare_text_files(file1, file2, output_file):
    """
    Compares two text files using difflib and saves the differences to an output file only if there are differences.

    Parameters:
    -----------
    - file1 : str
        Path to the first file to compare.
    - file2 : str
        Path to the second file to compare.
    - output_file : str
        Path to the output file where the differences will be saved.

    Returns:
    --------
    int
        The number of lines with differences.
    """
    # Read the content of the files
    with open(file1, 'r') as f1:
        text1 = f1.readlines()
    
    with open(file2, 'r') as f2:
        text2 = f2.readlines()
    
    # Use difflib to generate the differences
    diff = list(difflib.unified_diff(text1, text2, fromfile=file1, tofile=file2, n=0))
    
    # Filter out lines that start with '---', '+++' or '@@'
    filtered_diff = [line for line in diff if not (line.startswith('---') or line.startswith('+++') or line.startswith('@@'))]

    # Add a newline character to the end of each element in filtered_diff if it doesn't already have one, except for the last element
    for i in range(len(filtered_diff) - 1):  # Exclude the last element
        if not filtered_diff[i].endswith('\n'):
            filtered_diff[i] += '\n'
            
    # Count the number of lines with differences
    diff_lines = len(filtered_diff)

    # Only save the differences if there are any
    if diff_lines > 0:
        with open(output_file, 'w') as output:
            output.writelines(filtered_diff)
    
    return diff_lines


def compare_folders_and_save_diffs(folder1, folder2, diff_folder):
    """
    Compares the files in two folders and saves the differences in a third folder.
    Returns a DataFrame with the name of the diff file and the number of lines with differences.

    Parameters:
    -----------
    - folder1 : str
        Path to the first folder.
    - folder2 : str
        Path to the second folder.
    - diff_folder : str
        Path to the folder where the diff files will be saved.

    Returns:
    --------
    pandas.DataFrame
        A DataFrame with columns 'file_name', 'diff_file', 'diff_lines', and 'file_exists'.
    """

    # Create the diff folder if it doesn't exist
    if not os.path.exists(diff_folder):
        os.makedirs(diff_folder)
    
    # Get the list of text files in the first folder
    files1 = [f for f in os.listdir(folder1) if os.path.isfile(os.path.join(folder1, f))]

    # List to store the diff results
    diffs_data = []
    
    # Compare each file in the first folder with the corresponding file in the second folder
    for file_name in files1:
        file1_path = os.path.join(folder1, file_name)
        file2_path = os.path.join(folder2, file_name)

        # Check if the file also exists in the second folder
        if os.path.exists(file2_path):
            # Name of the diff file
            diff_file_name = f"diff_{os.path.splitext(file_name)[0]}.txt"
            diff_file_path = os.path.join(diff_folder, diff_file_name)
            
            # Call the comparison function and get the number of lines with differences
            diff_lines = compare_text_files(file1_path, file2_path, diff_file_path)
            
            # Add the result to the DataFrame 
            if diff_lines > 0:
                diffs_data.append({
                    'file_name': file_name,
                    'diff_file': diff_file_path,
                    'diff_lines': diff_lines,
                    'file_exists': True
                })
            else:
                diffs_data.append({
                    'file_name': file_name,
                    'diff_file': '',
                    'diff_lines': 0,
                    'file_exists': True
                })
        else:
            diffs_data.append({
                'file_name': file_name,
                'diff_file': '',
                'diff_lines': 0,
                'file_exists': False
            })
    
    # Create a DataFrame with the results
    diffs_df = pd.DataFrame(diffs_data, columns=['file_name', 'diff_file', 'diff_lines', 'file_exists'])

    return diffs_df


def generate_excel_from_diffs(folder1, folder2, diff_folder):
    """
    Genera un archivo Excel con las diferencias encontradas entre los archivos de dos carpetas.
    
    Parámetros:
    - folder1: Ruta a la primera carpeta.
    - folder2: Ruta a la segunda carpeta.
    - diff_folder: Ruta a la carpeta donde se guardarán los archivos de diferencias y el archivo Excel.
    
    El archivo Excel se guarda en 'diff_folder' con el nombre 'diff.xlsx'.
    """
    # Llama a la función para comparar las carpetas y obtener el dataframe
    diffs_df = compare_folders_and_save_diffs(folder1, folder2, diff_folder)
    
    # Solo procedemos si hay diferencias
    if not diffs_df.empty:
        # Crear un nuevo Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Differences"

        # Escribir la cabecera en el archivo Excel
        header = ['table_name', 'file_name', 'diff_file', 'diff_lines', 'file_exists']
        ws.append(header)

        # Formatear las celdas de cabecera
        for col_num, column_title in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col_num)
            format_header_cell(cell)  # Formatear la celda de cabecera
        
        # Iterar a través del DataFrame y extraer la información
        for index, row in diffs_df.iterrows():
            file = row['file_name']
            diff_file = row['diff_file']
            diff_lines = row['diff_lines']
            file_exists = row['file_exists']
            
            # Obtener el nombre de la tabla desde el nombre del fichero
            # table_name_match = re.search(r"diff_(.+?)__", diff_file)
            table_name_match = re.match(r"(.+?)__", file)
            table_name = table_name_match.group(1) if table_name_match else "Unknown"

            # Obtener solo el nombre del archivo para mostrar
            file_name = os.path.basename(diff_file)

            # Escribir la fila en el archivo Excel
            ws.append([table_name, file, diff_file, diff_lines, file_exists])

            # Configurar el hipervínculo en la celda 'diff_file'
            diff_file_cell = ws.cell(row=index + 2, column=3, value=file_name)  # +2 debido a la cabecera y la base 1 de openpyxl
            diff_file_cell.hyperlink = diff_file  # Establecer el hipervínculo a la ruta completa del archivo
            diff_file_cell.style = "Hyperlink"  # Estilo de hipervínculo para visualización


        # Inmovilizar la primera fila (cabecera)
        ws.freeze_panes = ws['A2']        
        
        # Auto-size columns 
        adjust_column_widths(ws)

        # Aplicar autofiltro a todas las columnas
        ws.auto_filter.ref = ws.dimensions

        # Guardar el archivo Excel en la carpeta diff_folder
        excel_path = os.path.join(diff_folder, 'diff.xlsx')
        wb.save(excel_path)


def compare_excel_dbinfo_files(file1, file2, output_file):
    """
    Compare two Excel files generated from get-dbinfo-metadata to find differences
    in column details for each table listed in the ALL_TABLES sheet. Saves the differences
    in a new Excel file with an additional column containing SQL statements to resolve
    the differences.

    Args:
        file1 (str): Path to the first Excel file.
        file2 (str): Path to the second Excel file.
        output_file (str): Path to the Excel file where differences will be saved.

    Returns:
        pd.DataFrame: A DataFrame containing the differences with columns:
                      'table_name', 'column_name', 'difference', and 'resolution_sql'.
    """
    # Load Excel files
    excel1 = pd.ExcelFile(file1)
    excel2 = pd.ExcelFile(file2)
    
    # Load ALL_TABLES sheets
    all_tables1 = pd.read_excel(excel1, sheet_name="ALL_TABLES")
    all_tables2 = pd.read_excel(excel2, sheet_name="ALL_TABLES")
    
    # Get the list of table names in each Excel file
    tables1 = set(all_tables1['table_name'].dropna().unique())
    tables2 = set(all_tables2['table_name'].dropna().unique())
    
    # Identify common tables and tables missing in each file
    common_tables = tables1.intersection(tables2)
    tables_only_in_file1 = tables1 - tables2
    tables_only_in_file2 = tables2 - tables1
    
    # Initialize list to store differences
    differences = []
    
    # Compare common tables
    for table_name in common_tables:
        # Load the table sheet from both Excel files
        df1 = pd.read_excel(excel1, sheet_name=table_name)
        df2 = pd.read_excel(excel2, sheet_name=table_name)
        
        # Compare the specific columns: column_name, data_type, data_length
        for _, row1 in df1.iterrows():
            column_name = row1['column_name']
            # Find the matching row in the second dataframe based on column_name
            row2 = df2[df2['column_name'] == column_name]
            
            if row2.empty:
                # Column is missing in the second file
                differences.append({
                    'table_name': table_name,
                    'column_name': column_name,
                    'difference': 'Column missing in second file',
                    'resolution_sql': f"ALTER TABLE SGLOWNER.{table_name} ADD {column_name} {row1['data_type']}({row1['data_length']});"
                })
                continue
            
            row2 = row2.iloc[0]  # There should be only one matching row
            
            # Compare data_type
            if row1['data_type'] != row2['data_type']:
                differences.append({
                    'table_name': table_name,
                    'column_name': column_name,
                    'difference': f"data_type mismatch: {row1['data_type']} vs {row2['data_type']}",
                    'resolution_sql': f"ALTER TABLE SGLOWNER.{table_name} MODIFY {column_name} {row1['data_type']}({row1['data_length']});"
                })
            
            # Compare data_length
            if row1['data_length'] != row2['data_length']:
                differences.append({
                    'table_name': table_name,
                    'column_name': column_name,
                    'difference': f"data_length mismatch: {row1['data_length']} vs {row2['data_length']}",
                    'resolution_sql': f"ALTER TABLE SGLOWNER.{table_name} MODIFY {column_name} {row1['data_type']}({row1['data_length']});"
                })

        # Check for columns in df2 that are not in df1
        for _, row2 in df2.iterrows():
            column_name = row2['column_name']
            if column_name not in df1['column_name'].values:
                differences.append({
                    'table_name': table_name,
                    'column_name': column_name,
                    'difference': 'Column missing in first file',
                    'resolution_sql': f"ALTER TABLE SGLOWNER.{table_name} ADD {column_name} {row2['data_type']}({row2['data_length']});"
                })

    # Add entries for tables only present in one of the files
    for table_name in tables_only_in_file1:
        differences.append({
            'table_name': table_name,
            'column_name': None,
            'difference': 'Table missing in second file',
            'resolution_sql': None
        })
    
    for table_name in tables_only_in_file2:
        differences.append({
            'table_name': table_name,
            'column_name': None,
            'difference': 'Table missing in first file',
            'resolution_sql': None
        })
    
    # Convert the differences list to a DataFrame
    differences_df = pd.DataFrame(differences, columns=['table_name', 'column_name', 'difference', 'resolution_sql'])
    
    # Save the differences DataFrame to the specified output Excel file
    with pd.ExcelWriter(output_file) as writer:
        differences_df.to_excel(writer, sheet_name="Differences", index=False)
    
    return differences_df


def get_folder_files_info(folder_path, file_extension=None):
    """
    Generates a DataFrame with information about files in a specified folder,
    optionally filtering by file extension.

    Args:
        folder_path (str): The path to the folder where files will be listed.
        file_extension (str, optional): The extension of the files to search for. If None, all files are listed.

    Returns:
        pd.DataFrame: A DataFrame with columns 'file_name', 'modification_date', and 'file_size' for each file.
    """
    # Initialize list to store file information
    files_data = []

    # Iterate over files in the folder
    for file_name in os.listdir(folder_path):
        # Full path of the file
        file_path = os.path.join(folder_path, file_name)
        
        # Check if it's a file and if it matches the extension (if provided)
        if os.path.isfile(file_path) and (file_extension is None or file_name.endswith(file_extension)):
            # Get file details
            modification_date = datetime.fromtimestamp(os.path.getmtime(file_path))
            file_size = os.path.getsize(file_path)
            
            # Append file info to the list
            files_data.append({
                'file_name': file_name,
                'modification_date': modification_date,
                'file_size': file_size
            })

    # Create DataFrame from the list
    files_df = pd.DataFrame(files_data, columns=['file_name', 'modification_date', 'file_size'])
    
    return files_df


def compare_file_info(df1, df2, output_folder):
    """
    Compare two DataFrames containing file information and extract differences.
    Save the differences in an Excel file in the specified output folder.

    Args:
        df1 (pd.DataFrame): First DataFrame containing file information.
        df2 (pd.DataFrame): Second DataFrame containing file information.
        output_folder (str): Folder where the Excel file with differences will be saved.

    Returns:
        pd.DataFrame: A DataFrame containing the differences with columns:
                      'file_name', 'difference_type', 'modification_date_df1', 
                      'modification_date_df2', 'file_size_df1', 'file_size_df2'.
    """
    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)
    
    # Find unique and common files by merging dataframes on file_name
    merged_df = df1.merge(df2, on='file_name', how='outer', suffixes=('_df1', '_df2'), indicator=True)

    # Initialize list to store differences
    differences = []

    # Iterate over merged DataFrame to identify differences
    for _, row in merged_df.iterrows():
        file_name = row['file_name']
        
        if row['_merge'] == 'left_only':
            # File only in the first DataFrame
            differences.append({
                'file_name': file_name,
                'difference_type': 'File only in first folder',
                'modification_date_df1': row['modification_date_df1'],
                'modification_date_df2': None,
                'file_size_df1': row['file_size_df1'],
                'file_size_df2': None
            })
        elif row['_merge'] == 'right_only':
            # File only in the second DataFrame
            differences.append({
                'file_name': file_name,
                'difference_type': 'File only in second folder',
                'modification_date_df1': None,
                'modification_date_df2': row['modification_date_df2'],
                'file_size_df1': None,
                'file_size_df2': row['file_size_df2']
            })
        else:
            # File in both DataFrames, check for differences in modification date and file size
            date_diff = row['modification_date_df1'] != row['modification_date_df2']
            size_diff = row['file_size_df1'] != row['file_size_df2']
            
            if date_diff or size_diff:
                difference_type = []
                if date_diff:
                    difference_type.append("Modification date mismatch")
                if size_diff:
                    difference_type.append("File size mismatch")
                
                differences.append({
                    'file_name': file_name,
                    'difference_type': ", ".join(difference_type),
                    'modification_date_df1': row['modification_date_df1'],
                    'modification_date_df2': row['modification_date_df2'],
                    'file_size_df1': row['file_size_df1'],
                    'file_size_df2': row['file_size_df2']
                })

    # Convert the differences list to a DataFrame
    differences_df = pd.DataFrame(differences, columns=[
        'file_name', 'difference_type', 'modification_date_df1', 
        'modification_date_df2', 'file_size_df1', 'file_size_df2'
    ])

    # Save differences to an Excel file in the output folder
    output_file_path = os.path.join(output_folder, "file_differences.xlsx")
    with pd.ExcelWriter(output_file_path) as writer:
        differences_df.to_excel(writer, sheet_name="Differences", index=False)
    
    return differences_df








